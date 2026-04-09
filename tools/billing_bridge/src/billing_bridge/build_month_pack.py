from __future__ import annotations

import argparse
import json
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ALIASES = {
    "Richard Perez": "Rich Perez",
    "Rich Perez": "Rich Perez",
    "Christopher Cummings": "Chris Cummings",
    "Chris Cummings": "Chris Cummings",
}


def canon(name: object) -> Optional[str]:
    if name is None:
        return None
    raw = str(name).strip()
    return ALIASES.get(raw, raw)


def parse_excel_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return pd.to_datetime(value).date()
        except Exception:
            return None
    return None


def parse_time(value: object) -> object:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, str):
        token = value.strip()
        if token.upper() in {"PTO", "OUT SICK", "N/A", "VACATION", "SICK"}:
            return token.upper()
        try:
            return pd.to_datetime(token).time()
        except Exception:
            return token
    return value


def hours_between(clock_in: object, clock_out: object) -> Optional[float]:
    if not isinstance(clock_in, time) or not isinstance(clock_out, time):
        return None
    start = datetime.combine(date(2026, 1, 1), clock_in)
    end = datetime.combine(date(2026, 1, 1), clock_out)
    if end < start:
        end += timedelta(days=1)
    return round((end - start).total_seconds() / 3600, 10)


def month_labels(year: int, month: int) -> tuple[str, str]:
    month_full = datetime(year, month, 1).strftime("%B")
    month_short = datetime(year, month, 1).strftime("%b")
    return month_full, f"{month_short} {str(year)[2:]}"


def extract_event_log(task_tracker: Path, year: int, month: int) -> pd.DataFrame:
    wb = load_workbook(task_tracker, data_only=True)
    ws = wb["Event Log"]
    headers = [ws.cell(5, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for r in range(6, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        work_date = parse_excel_date(row.get("Event Date"))
        if not work_date or work_date.year != year or work_date.month != month:
            continue
        rows.append(
            {
                "row": r,
                "date": work_date,
                "name": canon(row.get("Person")),
                "site": row.get("Site"),
                "workstream": row.get("Workstream"),
                "task_category": row.get("Task Category"),
                "evidence_source": row.get("Evidence Source"),
                "notes": row.get("Notes"),
                "actual_billed_hours": row.get("Actual Billed Hours"),
                "suggested_hours": row.get("Suggested Hours"),
            }
        )
    return pd.DataFrame(rows)


def extract_attendance(roster_path: Path, month_full: str, year: int) -> pd.DataFrame:
    wb = load_workbook(roster_path, data_only=False)
    wb_values = load_workbook(roster_path, data_only=True)
    ws = wb[f"Live - {month_full} {year}"]
    ws_values = wb_values[f"Live - {month_full} {year}"]
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

    rows: list[dict] = []
    for r in range(3, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name:
            continue
        name = canon(raw_name)
        project = ws_values.cell(r, 2).value or ws.cell(r, 2).value or ""
        for c in range(3, ws.max_column + 1, 2):
            header = headers[c - 1]
            if not header:
                continue
            date_text = str(header).split(" - ")[0]
            try:
                work_date = pd.to_datetime(f"{date_text} {year}").date()
            except Exception:
                continue
            clock_in = parse_time(ws_values.cell(r, c).value)
            clock_out = parse_time(ws_values.cell(r, c + 1).value if c + 1 <= ws.max_column else None)
            if clock_in is None and clock_out is None:
                continue
            rows.append(
                {
                    "date": work_date,
                    "name": name,
                    "project": project,
                    "clock_in": clock_in,
                    "clock_out": clock_out,
                    "hours": hours_between(clock_in, clock_out),
                    "source_ref": f"'Live - {month_full} {year}'!{get_column_letter(c)}{r}:{get_column_letter(c + 1)}{r}",
                }
            )
    return pd.DataFrame(rows)


def build_outputs(task_df: pd.DataFrame, attendance_df: pd.DataFrame, cfg: dict) -> tuple[pd.DataFrame, ...]:
    billing_rows: list[dict] = []
    field_rows: list[dict] = []
    experience_rows: list[dict] = []
    assignment_rows: list[dict] = []
    exception_rows: list[dict] = []

    for _, ev in task_df.sort_values(["date", "name"]).iterrows():
        matched = attendance_df[(attendance_df["date"] == ev["date"]) & (attendance_df["name"] == ev["name"])]
        field_rows.append(
            {
                "Date": ev["date"],
                "Staff Name": ev["name"],
                "Site": ev["site"],
                "Workstream": ev["workstream"],
                "Task Category": ev["task_category"],
                "Operational Insight": ev["evidence_source"],
                "Evidence Detail": ev["notes"],
                "Attendance Matched": "Yes" if not matched.empty else "No",
                "Attendance Source Ref": matched.iloc[0]["source_ref"] if not matched.empty else "",
                "Task Tracker Ref": f"Event Log!{ev['row']}",
            }
        )
        if ev["task_category"] or ev["workstream"] or ev["evidence_source"]:
            experience_rows.append(
                {
                    "Date": ev["date"],
                    "Staff Name": ev["name"],
                    "Experience Type": ev["task_category"] or "Unclassified",
                    "Skill Family": ev["workstream"] or "Unclassified",
                    "Site": ev["site"],
                    "Experience Summary": ev["evidence_source"] or "",
                    "Task Tracker Ref": f"Event Log!{ev['row']}",
                    "Attendance Matched": "Yes" if not matched.empty else "No",
                }
            )

    for _, att in attendance_df.sort_values(["date", "name"]).iterrows():
        evs = task_df[(task_df["date"] == att["date"]) & (task_df["name"] == att["name"])]
        if evs.empty:
            exception_rows.append(
                {
                    "Issue Type": "Attendance without task evidence",
                    "Date": att["date"],
                    "Staff Name": att["name"],
                    "Attendance Hours": att["hours"],
                    "Task Evidence Rows": 0,
                    "Source Ref": att["source_ref"],
                    "Details": "Attendance exists but no matching Event Log evidence was found for this tech/date.",
                    "Resolution Guidance": "Review the task tracker week tabs or create a manual override.",
                }
            )
            continue

        evs = evs[(evs["task_category"].notna()) | (evs["workstream"].notna()) | (evs["evidence_source"].notna())]
        if evs.empty:
            exception_rows.append(
                {
                    "Issue Type": "Task evidence missing category/workstream",
                    "Date": att["date"],
                    "Staff Name": att["name"],
                    "Attendance Hours": att["hours"],
                    "Task Evidence Rows": 0,
                    "Source Ref": att["source_ref"],
                    "Details": "Event Log rows exist, but categories/workstreams are blank, so billing allocation was not auto-created.",
                    "Resolution Guidance": "Backfill the Event Log or use a reviewed override.",
                }
            )
            continue

        if att["hours"] is None:
            exception_rows.append(
                {
                    "Issue Type": "Attendance code / no computable hours",
                    "Date": att["date"],
                    "Staff Name": att["name"],
                    "Attendance Hours": None,
                    "Task Evidence Rows": len(evs),
                    "Source Ref": att["source_ref"],
                    "Details": "Attendance exists but no computable time window was available.",
                    "Resolution Guidance": "Review the raw punches manually.",
                }
            )
            continue

        split_count = len(evs)
        allocated_hours = round(att["hours"] / split_count, 10)
        for _, ev in evs.iterrows():
            task_category = ev["task_category"] or ""
            workstream = ev["workstream"] or ""
            delivery = any(term in task_category.lower() for term in cfg["delivery_keywords"])
            bucket = cfg["delivery_bucket"] if delivery else cfg["default_bucket"]
            manager_assignment = cfg["manager_assignment_delivery"] if delivery else cfg["manager_assignment_default"]

            notes: list[str] = []
            if task_category:
                notes.append(task_category)
            if workstream:
                notes.append(workstream)
            if ev["evidence_source"]:
                notes.append(f"Evidence: {ev['evidence_source']}")
            if split_count > 1:
                notes.append(f"Split rule: equal split across {split_count} event rows for same tech/date")

            billing_rows.append(
                {
                    "Staff Name": ev["name"],
                    "Date": ev["date"],
                    "Worked Project": att["project"] or "Neuron Deployments",
                    "Billing Bucket": bucket,
                    "Clock In": att["clock_in"],
                    "Clock Out": att["clock_out"],
                    "Hours": allocated_hours,
                    "Billable Flag": "Billable",
                    "Source Ref": att["source_ref"],
                    "Notes": " | ".join(notes),
                    "Manager Assignment": manager_assignment,
                }
            )
            assignment_rows.append(
                {
                    "Date": ev["date"],
                    "Staff Name": ev["name"],
                    "Worked Project": att["project"] or "Neuron Deployments",
                    "Billing Bucket": bucket,
                    "Manager Assignment": manager_assignment,
                    "Attendance Hours": att["hours"],
                    "Allocated Hours": allocated_hours,
                    "Split Rule Applied": "Yes" if split_count > 1 else "No",
                    "Task Category": task_category,
                    "Workstream": workstream,
                    "Operational Insight": ev["evidence_source"],
                    "Task Tracker Ref": f"Event Log!{ev['row']}",
                    "Attendance Source Ref": att["source_ref"],
                }
            )

    for _, ev in task_df.sort_values(["date", "name"]).iterrows():
        matched = attendance_df[(attendance_df["date"] == ev["date"]) & (attendance_df["name"] == ev["name"])]
        if not matched.empty:
            continue
        exception_rows.append(
            {
                "Issue Type": "Task evidence without attendance",
                "Date": ev["date"],
                "Staff Name": ev["name"],
                "Attendance Hours": None,
                "Task Evidence Rows": 1,
                "Source Ref": f"Event Log!{ev['row']}",
                "Details": "Event Log evidence exists but no matching attendance row was found.",
                "Resolution Guidance": "Review roster punches, attendance bridge, or manual override.",
            }
        )

    return (
        pd.DataFrame(billing_rows).sort_values(["Date", "Staff Name", "Billing Bucket"]),
        pd.DataFrame(field_rows).sort_values(["Date", "Staff Name"]),
        pd.DataFrame(experience_rows).sort_values(["Date", "Staff Name"]),
        pd.DataFrame(assignment_rows).sort_values(["Date", "Staff Name"]),
        pd.DataFrame(exception_rows).sort_values(["Date", "Staff Name", "Issue Type"]),
    )


def clear_values(ws, start_row: int, max_col: int) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def style_title(ws, title: str, subtitle: str, end_col: int) -> None:
    dark_fill = PatternFill("solid", fgColor="1F1F1F")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].fill = dark_fill
    ws["A1"].font = white_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].fill = sub_fill
    ws["A2"].alignment = Alignment(wrap_text=True)


def write_table(ws, start_row: int, df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="244062")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")

    headers = list(df.columns) if not df.empty else ["Note"]
    rows = df.to_dict("records") if not df.empty else [{"Note": "No rows generated."}]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, i)
        cell.value = header
        cell.fill = header_fill
        cell.font = white_font
        cell.border = Border(bottom=Border(bottom=thin).bottom)
    for ridx, row in enumerate(rows, start=start_row + 1):
        for cidx, header in enumerate(headers, start=1):
            cell = ws.cell(ridx, cidx)
            cell.value = row.get(header)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = f"A{start_row + 1}"


def add_or_replace_sheet(wb, title: str):
    if title in wb.sheetnames:
        wb.remove(wb[title])
    return wb.create_sheet(title)


def build_workbook(task_tracker: Path, roster: Path, month: str, components: set[str], output: Path) -> None:
    year, month_num = [int(part) for part in month.split("-")]
    month_full, month_short_label = month_labels(year, month_num)
    cfg = json.loads((Path(__file__).resolve().parent.parent / "config" / "assignment_map.json").read_text())

    task_df = extract_event_log(task_tracker, year, month_num)
    attendance_df = extract_attendance(roster, month_full, year)
    billing_df, field_df, experience_df, assignment_df, exceptions_df = build_outputs(task_df, attendance_df, cfg)

    wb = load_workbook(roster)
    month_short = datetime(year, month_num, 1).strftime("%b")

    if "all" in components or "billing" in components:
        template_name = "Billing Detail - Mar 2026"
        if template_name not in wb.sheetnames:
            raise ValueError(f"Template sheet '{template_name}' not found in workbook. Available: {wb.sheetnames}")
        source = wb[template_name]
        if f"Billing Detail - {month_full} {year}" in wb.sheetnames:
            wb.remove(wb[f"Billing Detail - {month_full} {year}"])
        detail_ws = wb.copy_worksheet(source)
        detail_ws.title = f"Billing Detail - {month_full} {year}"
        detail_ws["A1"] = f"{month_short} {year} - Billing Detail (Northwell Neurons + Delivery / Transport / Disposal)"
        detail_ws["A2"] = "Evidence-first billing rows. Attendance caps hours; same-day multi-event splits use an equal split rule and are flagged in companion tabs."
        clear_values(detail_ws, 4, 10)
        for idx, row in billing_df.reset_index(drop=True).iterrows():
            r = 4 + idx
            detail_ws.cell(r, 1).value = row["Staff Name"]
            detail_ws.cell(r, 2).value = datetime.combine(row["Date"], time())
            detail_ws.cell(r, 3).value = row["Worked Project"]
            detail_ws.cell(r, 4).value = row["Billing Bucket"]
            detail_ws.cell(r, 5).value = row["Clock In"] if isinstance(row["Clock In"], time) else None
            detail_ws.cell(r, 6).value = row["Clock Out"] if isinstance(row["Clock Out"], time) else None
            detail_ws.cell(r, 7).value = row["Hours"]
            detail_ws.cell(r, 8).value = row["Billable Flag"]
            detail_ws.cell(r, 9).value = row["Source Ref"]
            detail_ws.cell(r, 10).value = row["Notes"]

    if "all" in components or "billing-summary" in components or "billing" in components:
        source = wb["Billing Summary - Mar 2026"]
        if f"Billing Summary - {month_full} {year}" in wb.sheetnames:
            wb.remove(wb[f"Billing Summary - {month_full} {year}"])
        summary_ws = wb.copy_worksheet(source)
        summary_ws.title = f"Billing Summary - {month_full} {year}"
        summary_ws["A1"] = f"{month_short} {year} - Northwell Billing Summary"
        summary_ws["A2"] = "Monthly rollup driven by Billing Detail. Use Field Insights and Bridge Exceptions for the fuller operational picture."
        clear_values(summary_ws, 5, 8)

        neurons = billing_df[billing_df["Billing Bucket"] == "Neurons"]
        delivery = billing_df[billing_df["Billing Bucket"] == "Delivery / Transport / Disposal"]

        summary_ws["A4"] = "Billing Bucket"
        summary_ws["B4"] = "Tech Count"
        summary_ws["C4"] = "Worked Rows"
        summary_ws["D4"] = "Billable Hours"
        summary_ws["A5"] = "Neurons"
        summary_ws["B5"] = int(neurons["Staff Name"].nunique())
        summary_ws["C5"] = int(len(neurons))
        summary_ws["D5"] = float(round(neurons["Hours"].sum(), 2))
        summary_ws["A6"] = "Delivery / Transport / Disposal"
        summary_ws["B6"] = int(delivery["Staff Name"].nunique())
        summary_ws["C6"] = int(len(delivery))
        summary_ws["D6"] = float(round(delivery["Hours"].sum(), 2))
        summary_ws["A7"] = "Combined Northwell billing"
        summary_ws["B7"] = int(billing_df["Staff Name"].nunique())
        summary_ws["C7"] = int(len(billing_df))
        summary_ws["D7"] = float(round(billing_df["Hours"].sum(), 2))
        summary_ws["A9"] = "Equal split rule used only when one attendance row matches multiple same-day Event Log rows without explicit billed-hour weights."
        summary_ws["A10"] = "Neurons by Tech"
        summary_ws["F10"] = "Delivery / Transport / Disposal by Tech"
        summary_ws["A11"] = "Staff Name"
        summary_ws["B11"] = "Worked Rows"
        summary_ws["C11"] = "Billable Hours"
        summary_ws["D11"] = "Billing Note"
        summary_ws["F11"] = "Staff Name"
        summary_ws["G11"] = "Worked Rows"
        summary_ws["H11"] = "Billable Hours"

        neuron_tech = (
            neurons.groupby("Staff Name", as_index=False)
            .agg(**{"Worked Rows": ("Date", "count"), "Billable Hours": ("Hours", "sum")})
            .sort_values(["Billable Hours", "Staff Name"], ascending=[False, True])
        )
        delivery_tech = (
            delivery.groupby("Staff Name", as_index=False)
            .agg(**{"Worked Rows": ("Date", "count"), "Billable Hours": ("Hours", "sum")})
            .sort_values(["Billable Hours", "Staff Name"], ascending=[False, True])
        )

        row_num = 12
        for _, row in neuron_tech.iterrows():
            summary_ws.cell(row_num, 1).value = row["Staff Name"]
            summary_ws.cell(row_num, 2).value = int(row["Worked Rows"])
            summary_ws.cell(row_num, 3).value = float(round(row["Billable Hours"], 2))
            summary_ws.cell(row_num, 4).value = "Bill to Northwell - Neurons"
            row_num += 1

        row_num = 12
        for _, row in delivery_tech.iterrows():
            summary_ws.cell(row_num, 6).value = row["Staff Name"]
            summary_ws.cell(row_num, 7).value = int(row["Worked Rows"])
            summary_ws.cell(row_num, 8).value = float(round(row["Billable Hours"], 2))
            row_num += 1
        summary_ws.cell(row_num, 6).value = "Transport Total"
        summary_ws.cell(row_num, 7).value = int(len(delivery))
        summary_ws.cell(row_num, 8).value = float(round(delivery["Hours"].sum(), 2))

    if "all" in components or "bonita" in components:
        source = wb["Bonitas Tracker - Mar 26"]
        target_title = f"Bonitas Tracker - {month_short} {str(year)[2:]}"
        if target_title in wb.sheetnames:
            wb.remove(wb[target_title])
        bonita_ws = wb.copy_worksheet(source)
        bonita_ws.title = target_title
        clear_values(bonita_ws, 3, 7)
        row_num = 3
        for work_date, day_rows in billing_df.sort_values(["Date", "Staff Name", "Billing Bucket"]).groupby("Date"):
            for _, row in day_rows.iterrows():
                bonita_ws.cell(row_num, 1).value = datetime.combine(work_date, time())
                bonita_ws.cell(row_num, 2).value = row["Staff Name"]
                bonita_ws.cell(row_num, 3).value = row["Clock In"] if isinstance(row["Clock In"], time) else None
                bonita_ws.cell(row_num, 4).value = row["Clock Out"] if isinstance(row["Clock Out"], time) else None
                bonita_ws.cell(row_num, 5).value = row["Hours"]
                bonita_ws.cell(row_num, 6).value = cfg["manager_project"]
                bonita_ws.cell(row_num, 7).value = row["Manager Assignment"]
                row_num += 1

    if "all" in components or "field-insights" in components:
        ws = add_or_replace_sheet(wb, f"{month_short} {str(year)[2:]} - Field Insights")
        style_title(ws, f"{month_short} {str(year)[2:]} - Field Insights", "Operational insights derived from Event Log rows. Attendance match is shown, but this tab is not the billing source.", max(1, len(field_df.columns)))
        write_table(ws, 4, field_df)

    if "all" in components or "experience-ledger" in components:
        ws = add_or_replace_sheet(wb, f"{month_short} {str(year)[2:]} - Experience Ledger")
        style_title(ws, f"{month_short} {str(year)[2:]} - Experience Ledger", "Resume-protecting monthly ledger of demonstrated work categories and skill families.", max(1, len(experience_df.columns)))
        write_table(ws, 4, experience_df)

    if "all" in components or "assignment-evidence" in components:
        ws = add_or_replace_sheet(wb, f"{month_short} {str(year)[2:]} - Assignment Evidence")
        style_title(ws, f"{month_short} {str(year)[2:]} - Assignment Evidence", "One row per billed allocation, tying outward billing back to attendance and Event Log evidence.", max(1, len(assignment_df.columns)))
        write_table(ws, 4, assignment_df)

    if "all" in components or "bridge-exceptions" in components:
        ws = add_or_replace_sheet(wb, f"{month_short} {str(year)[2:]} - Bridge Exceptions")
        style_title(ws, f"{month_short} {str(year)[2:]} - Bridge Exceptions", "Rows that require review instead of invention. Attendance-only days and evidence-only days land here.", max(1, len(exceptions_df.columns)))
        write_table(ws, 4, exceptions_df)

    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--task-tracker", required=True)
    parser.add_argument("--roster", required=True)
    parser.add_argument("--month", required=True, help="YYYY-MM")
    parser.add_argument("--output", required=True)
    parser.add_argument(
        "--components",
        nargs="+",
        default=["all"],
        choices=[
            "all",
            "billing",
            "billing-summary",
            "bonita",
            "field-insights",
            "experience-ledger",
            "assignment-evidence",
            "bridge-exceptions",
        ],
    )
    args = parser.parse_args()

    build_workbook(
        task_tracker=Path(args.task_tracker),
        roster=Path(args.roster),
        month=args.month,
        components=set(args.components),
        output=Path(args.output),
    )


if __name__ == "__main__":
    main()