from __future__ import annotations

from pathlib import Path

import pandas as pd


MANAGER_COLUMNS = ["TECH", "START", "END", "TOTAL", "PROJECT", "ASSIGNMENT"]


def build_manager_rows(allocations: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame(
        {
            "TECH": allocations["canonical_name"],
            "START": allocations["clock_in"],
            "END": allocations["clock_out"],
            "TOTAL": allocations["allocated_hours"],
            "PROJECT": allocations["outward_project"],
            "ASSIGNMENT": allocations["outward_assignment"],
        }
    )
    return out[MANAGER_COLUMNS]


def replace_month_sheet(
    workbook_path: str | Path,
    month_sheet: str,
    manager_rows: pd.DataFrame,
    output_path: str | Path,
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    ws = wb[month_sheet]

    start_row = 7  # adjust after inspecting real manager workbook conventions
    # clear existing content region conservatively
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=2, max_col=7):
        for cell in row:
            cell.value = None

    for idx, row in enumerate(manager_rows.itertuples(index=False), start=start_row):
        ws.cell(idx, 2).value = row.TECH
        ws.cell(idx, 3).value = row.START
        ws.cell(idx, 4).value = row.END
        ws.cell(idx, 5).value = row.TOTAL
        ws.cell(idx, 6).value = row.PROJECT
        ws.cell(idx, 7).value = row.ASSIGNMENT

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
